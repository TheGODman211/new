import os

import openpyxl
import openpyxl as px
import glob

import string
from openpyxl.reader.excel import load_workbook
import pyexcel as p
import datetime

#------PARSE XLS TO WLSX-------
# p.save_book_as(file_name='D:\excel\SUPERVISION36.xls', dest_file_name='D:/excel/Updated.xlsx')



# logs = [log for log in glob('D:\excel\*') if not os.path.isdir(log)]
# -----Pass a list to this function to check for maximum number-----
# def max_check(p):
#     max_val = p[1]
#     for check in p[1:]:
#         if check > max_val:
#             max_val = check
#     return max_val
#
#
# ------- Pass a list to this function to check for minimum number
# def min_check(o):
#     min_val = o[1]
#     for check in o[1:]:
#         if check < min_val:
#             min_val = check
#     return min_val


# List


# Printing Values

# print("Maximum of the list", max_check(my_list))
# print("Minimum of the list", min_check(my_list))

# ------RETRIEVE EXCEL FILE Paths------
# gets the excel file paths recursively and saves in a list
excel_files = glob.glob('D:/excel/*', recursive=True)
# print(excel_files)
# print(len(excel_files))
value = []

#------CONVERT XLS TO XLSX------
for file in excel_files:
    if file[-3:] == 'xls':
        filename = file[0:-4]
        new = filename + '.xlsx'
        p.save_book_as(file_name=file, dest_file_name=new)
    #     wb = load_workbook(new, data_only=True)
    #     sheet_obj = wb.active


#------INITIALISE A LIST OF APLHABETS -----
test_list = []
test_list = list(string.ascii_uppercase)
x = 68
i = 4
r = 0

#--------CREATE NEW WORKBOOK----
wa = openpyxl.Workbook()
sheet_now = wa.active

sheet_now['A2'].value = 'NAME'
sheet_now['B2'].value = 'TIME IN'
sheet_now['C2'].value = 'TIME OUT'

#-----SEARCH FOR ALL EXCEL FILES BEGINNING WITH S IN THE FOLDER----
new_excel_files = glob.glob('D:/excel/S*.xlsx', recursive=True)
print(new_excel_files)

#------LOAD ALL WORKBOOKS AND ITERATE THROUGH THEM-----
for new_file in new_excel_files:
    wb = load_workbook(new_file, data_only=True)
   # wb.active = wb['SUPERVISION 3rd-6th January']
    sheet_obj = wb.active
    print(wb.active)


    for indice, row in enumerate(sheet_obj.iter_rows(min_row=2, max_row=33, min_col=7, max_col=7, values_only=True),
                                  start=2):
        # if x < indice:
        #     # test_list[i] = []
        #     # exec('%s = %' % (test_list[i], []))
        #     # locals()[test_list[i]] = []
        #     i += 1
        a,d, e, f,g =[],[],[],[],[]
        for index, names in enumerate(sheet_obj.iter_rows(min_row=2, max_col=1, min_col=1), start=2):
            # print(datetime.date.isoformat(sheet_obj.cell(row=index, column=5).value))
            # print(sheet_obj.cell(row=indice, column=7).value)
            # y= datetime.date(2023, 1, 4)
            # print(datetime.date.isoformat(y))

            if sheet_obj.cell(row=index, column=1).value == sheet_obj.cell(row=indice,column=7).value and datetime.date.isoformat(
                    sheet_obj.cell(row=index, column=5).value) ==  datetime.date.isoformat(sheet_obj.cell(row=2, column=8).value):
                a.append(sheet_obj.cell(row=index, column=6).value)
            elif sheet_obj.cell(row=index, column=1).value == sheet_obj.cell(row=indice,
                                                                           column=7).value and datetime.date.isoformat(
                    sheet_obj.cell(row=index, column=5).value) == datetime.date.isoformat(sheet_obj.cell(row=3, column=8).value):
                d.append(sheet_obj.cell(row=index, column=6).value)
            elif sheet_obj.cell(row=index, column=1).value == sheet_obj.cell(row=indice,
                                                                           column=7).value and datetime.date.isoformat(
                    sheet_obj.cell(row=index, column=5).value) == datetime.date.isoformat(sheet_obj.cell(row=4, column=8).value):
                e.append(sheet_obj.cell(row=index, column=6).value)
            elif sheet_obj.cell(row=index, column=1).value == sheet_obj.cell(row=indice,
                                                                           column=7).value and datetime.date.isoformat(
                    sheet_obj.cell(row=index, column=5).value) == datetime.date.isoformat(sheet_obj.cell(row=5, column=8).value):
                f.append(sheet_obj.cell(row=index, column=6).value)
            elif sheet_obj.cell(row=index, column=1).value == sheet_obj.cell(row=indice,
                                                                           column=7).value and datetime.date.isoformat(
                    sheet_obj.cell(row=index, column=5).value) == datetime.date.isoformat(sheet_obj.cell(row=6, column=8).value):
                g.append(sheet_obj.cell(row=index, column=6).value)

        a.insert(0, sheet_obj.cell(row=indice, column=7).value)
        d.insert(0, sheet_obj.cell(row=indice, column=7).value)
        e.insert(0, sheet_obj.cell(row=indice, column=7).value)
        f.insert(0, sheet_obj.cell(row=indice, column=7).value)
        g.insert(0, sheet_obj.cell(row=indice, column=7).value)

        for place, obj in enumerate(a[1:]):   # remove time values less than 5am
            if obj < datetime.time(5, 0, 0):
                a.remove(obj)
        if len(a) == 1:
            b= 'ABSENT'; c = 'ABSENT'
        else:
            b = min(a[1:])
            c = max(a[1:])


        for place, obj in enumerate(d[1:]):   # remove time values less than 5am
            if obj < datetime.time(5, 0, 0):
                d.remove(obj)
        if len(d) == 1:
            b1 = 'ABSENT';
            c1 = 'ABSENT'
        else:
            b1 = min(d[1:])
            c1 = max(d[1:])

        for place, obj in enumerate(e[1:]):   # remove time values less than 5am
            if obj < datetime.time(5, 0, 0):
                e.remove(obj)
        if len(e) == 1:
            b2 = 'ABSENT';
            c2 = 'ABSENT'
        else:
            b2 = min(e[1:])
            c2 = max(e[1:])

        for place, obj in enumerate(f[1:]):   # remove time values less than 5am
            if obj < datetime.time(5, 0, 0):
                f.remove(obj)
        if len(f) == 1:
            b3 = 'ABSENT';
            c3 = 'ABSENT'
        else:
            b3 = min(f[1:])
            c3 = max(f[1:])
        for place, obj in enumerate(g[1:]):   # remove time values less than 5am
            if obj < datetime.time(5, 0, 0):
                a.remove(obj)
        if len(g) == 1:
            b4= 'ABSENT'; c4 = 'ABSENT'
        else:
            b4 = min(g[1:])
            c4 = max(g[1:])
        #sheet_now.cell(row=1, column=2).value = sheet_obj.cell(row=2, column=8).value
        sheet_now.cell(row=i, column=1).value = sheet_obj.cell(row=indice, column=7).value
        sheet_now.cell(row=i, column=2).value = b
        sheet_now.cell(row=i, column=3).value = c
       # sheet_now.cell(row=1, column=6).value = sheet_obj.cell(row=3, column=8).value
        sheet_now.cell(row=i, column=6).value = b1
        sheet_now.cell(row=i, column=7).value = c1
        #sheet_now.cell(row=1, column=10).value = sheet_obj.cell(row=4, column=8).value
        sheet_now.cell(row=i, column=10).value = b2
        sheet_now.cell(row=i, column=11).value = c2
        #sheet_now.cell(row=1, column=14).value = sheet_obj.cell(row=5, column=8).value
        sheet_now.cell(row=i, column=14).value = b3
        sheet_now.cell(row=i, column=15).value = c3
        #sheet_now.cell(row=1, column=18).value = sheet_obj.cell(row=6, column=8).value
        sheet_now.cell(row=i, column=18).value = b4
        sheet_now.cell(row=i, column=19).value = c4

        #wa.save('D:/excel/New.xlsx')
        #
        print(a)
        # print(b)
        # print(c)

        i +=1
    sheet_now.cell(row=r+1, column=2).value = sheet_obj.cell(row=2, column=8).value
    sheet_now.cell(row=r + 2, column=2).value = 'TIME IN'
    sheet_now.cell(row=r + 2, column=3).value = 'TIME OUT'
    sheet_now.cell(row=r+1, column=6).value = sheet_obj.cell(row=3, column=8).value
    sheet_now.cell(row=r + 2, column=6).value = 'TIME IN'
    sheet_now.cell(row=r + 2, column=7).value = 'TIME OUT'
    sheet_now.cell(row=r+1, column=10).value = sheet_obj.cell(row=4, column=8).value
    sheet_now.cell(row=r + 2, column=10).value = 'TIME IN'
    sheet_now.cell(row=r + 2, column=11).value = 'TIME OUT'
    sheet_now.cell(row=r+1, column=14).value = sheet_obj.cell(row=5, column=8).value
    sheet_now.cell(row=r + 2, column=14).value = 'TIME IN'
    sheet_now.cell(row=r + 2, column=15).value = 'TIME OUT'
    sheet_now.cell(row=r+1, column=18).value = sheet_obj.cell(row=6, column=8).value
    sheet_now.cell(row=r + 2, column=18).value = 'TIME IN'
    sheet_now.cell(row=r + 2, column=19).value = 'TIME OUT'

    print(sheet_obj.cell(row=2, column=8).value)
    print(sheet_obj.cell(row=3, column=8).value)
    print(sheet_obj.cell(row=4, column=8).value)
    print(sheet_obj.cell(row=5, column=8).value)
    print(sheet_obj.cell(row=6, column=8).value)
    r += 36
    i+=4
wa.save('D:/excel/Newt.xlsx')
        #     print(sheet_obj.cell(row=index, column=5).value,row,names)
        # elif names == row and sheet_obj.cell(row=index, column=5).value == datetime.date(2023, 1, 4):
        #     b.append(sheet_obj.cell(row=index, column=6).value)
        # elif names == row and sheet_obj.cell(row=index, column=5).value == datetime.date(2023, 1, 5):
        #     c.append(sheet_obj.cell(row=index, column=6).value)
        # elif names == row and sheet_obj.cell(row=index, column=5).value == datetime.date(2023, 1, 6):
        #     d.append(sheet_obj.cell(row=index, column=6).value)

# #-----RETRIEVE DATA IN SPECIFIC CELL----
# for excel_file in excel_files:
#     wb = load_workbook(excel_file, data_only=True)  # load workbook and cell values as data not formula
#     wb.active = wb[('SDR10iv')] #sets active sheet to SDR10iv
#     sheet_obj = wb.active
#     cell_obj = sheet_obj.cell(row=20, column=10)
#     value.append(cell_obj.value)
# print(value)
